import csv
import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.student import Student
from models.school_level import SchoolLevel
from services.student_service import StudentService


class StudentImportService:

    @staticmethod
    def clean_value(value):
        if value is None:
            return ""

        value = str(value).strip()

        if value.endswith(".0"):
            value = value[:-2]

        return value

    @staticmethod
    def normalize_alias(value):
        return (
            StudentImportService.clean_value(value)
            .upper()
            .replace(" ", "_")
            .replace("-", "_")
        )

    @staticmethod
    def normalize_money(value):
        value = StudentImportService.clean_value(value)

        if not value:
            return 0

        try:
            return float(
                value
                .replace("$", "")
                .replace(".", "")
                .replace(",", ".")
                .strip()
            )
        except Exception:
            return 0

    @staticmethod
    def normalize_row(row: dict):
        return {
            str(key).strip(): value
            for key, value in row.items()
            if key is not None
        }

    @staticmethod
    def read_csv(file_content: bytes):
        text = file_content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

        return [
            StudentImportService.normalize_row(row)
            for row in reader
        ]

    @staticmethod
    def read_xlsx(file_content: bytes):
        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            data_only=True
        )

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [
            StudentImportService.clean_value(header)
            for header in rows[0]
        ]

        data = []

        for row in rows[1:]:
            item = {}

            for index, header in enumerate(headers):
                item[header] = row[index] if index < len(row) else None

            data.append(
                StudentImportService.normalize_row(item)
            )

        return data

    @staticmethod
    def find_level_by_alias(
        db: Session,
        school_id: int,
        level_alias: str
    ):
        alias = StudentImportService.normalize_alias(level_alias)

        if not alias:
            return None

        return (
            db.query(SchoolLevel)
            .filter(
                SchoolLevel.school_id == school_id,
                SchoolLevel.alias == alias,
                SchoolLevel.is_active == True
            )
            .first()
        )

    @staticmethod
    def import_file(
        db: Session,
        school_id: int,
        file_content: bytes,
        filename: str
    ):
        filename = filename.lower()

        if filename.endswith(".csv"):
            rows = StudentImportService.read_csv(file_content)

        elif filename.endswith(".xlsx"):
            rows = StudentImportService.read_xlsx(file_content)

        else:
            return {
                "imported": 0,
                "skipped": 0,
                "errors": [
                    "Formato no permitido. Usá .csv o .xlsx"
                ]
            }

        imported = 0
        skipped = 0
        errors = []

        for index, row in enumerate(rows, start=2):

            full_name = StudentImportService.clean_value(
                row.get("full_name")
                or row.get("nombre")
                or row.get("Nombre")
                or ""
            )

            dni = StudentImportService.clean_value(
                row.get("dni")
                or row.get("DNI")
                or ""
            )

            guardian_dni = StudentImportService.clean_value(
                row.get("guardian_dni")
                or row.get("dni_tutor")
                or row.get("DNI Tutor")
                or row.get("dni tutor")
                or ""
            )

            if not full_name or not dni:
                skipped += 1
                errors.append(
                    f"Fila {index}: falta nombre o DNI"
                )
                continue

            try:
                clean_student_dni = StudentService.validate_dni(
                    dni,
                    "DNI del alumno"
                )

                clean_guardian_dni = None

                if guardian_dni:
                    clean_guardian_dni = StudentService.validate_dni(
                        guardian_dni,
                        "DNI del tutor"
                    )

                    if clean_guardian_dni == clean_student_dni:
                        raise ValueError(
                            "El DNI del tutor no puede ser igual al DNI del alumno."
                        )

            except Exception as e:
                skipped += 1
                errors.append(
                    f"Fila {index}: {str(e)}"
                )
                continue

            existing = db.query(Student).filter(
                Student.school_id == school_id,
                Student.dni == clean_student_dni
            ).first()

            if existing:
                skipped += 1
                errors.append(
                    f"Fila {index}: el DNI {clean_student_dni} ya existe en esta escuela"
                )
                continue

            level_alias = StudentImportService.clean_value(
                row.get("nivel_alias")
                or row.get("Nivel Alias")
                or row.get("alias_nivel")
                or row.get("level_alias")
                or ""
            )

            if not level_alias:
                skipped += 1
                errors.append(
                    f"Fila {index}: falta nivel_alias"
                )
                continue

            level = StudentImportService.find_level_by_alias(
                db=db,
                school_id=school_id,
                level_alias=level_alias
            )

            if not level:
                skipped += 1
                errors.append(
                    f"Fila {index}: no existe un nivel activo con alias '{level_alias}'. Crealo primero en Niveles."
                )
                continue

            monthly_fee = StudentImportService.normalize_money(
                row.get("monthly_fee")
                or row.get("cuota")
                or row.get("Cuota")
                or ""
            )

            if monthly_fee == 0:
                monthly_fee = float(level.monthly_fee or 0)

            student = Student(
                school_id=school_id,
                level_id=level.id,
                full_name=full_name,
                dni=clean_student_dni,
                email=StudentImportService.clean_value(
                    row.get("email")
                    or row.get("Email")
                    or ""
                ),
                course=StudentImportService.clean_value(
                    row.get("course")
                    or row.get("curso")
                    or row.get("Curso")
                    or ""
                ),
                division=StudentImportService.clean_value(
                    row.get("division")
                    or row.get("división")
                    or row.get("Division")
                    or ""
                ),
                guardian_name=StudentImportService.clean_value(
                    row.get("guardian_name")
                    or row.get("tutor")
                    or row.get("Tutor")
                    or ""
                ),
                guardian_dni=clean_guardian_dni,
                monthly_fee=monthly_fee,
                is_active=True
            )

            db.add(student)
            imported += 1

        db.commit()

        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors
        }