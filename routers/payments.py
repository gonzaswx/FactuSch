from datetime import date
from datetime import datetime
from datetime import timezone
import tempfile

from fastapi import APIRouter
from fastapi import Depends
from fastapi import Form
from fastapi import Request

from fastapi.responses import FileResponse
from fastapi.responses import HTMLResponse
from fastapi.responses import RedirectResponse

from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from database import get_db

from models.invoice import Invoice
from models.school import School
from models.school_settings import SchoolSettings
from models.student import Student

from routers.auth import require_school_access

from services.email_service import EmailService
from services.invoice_log_service import InvoiceLogService

from utils.templates import templates


router = APIRouter(
    prefix="/schools/{school_id}/payments",
    tags=["Payments"]
)


def parse_optional_int(value):
    if value is None or value == "":
        return None

    try:
        return int(value)
    except ValueError:
        return None


def get_invoice_number(invoice: Invoice):
    if invoice.full_invoice_number:
        return invoice.full_invoice_number

    return f"{int(invoice.point_of_sale or 0):04d}-{int(invoice.invoice_number or 0):08d}"


def build_payments_query(
    db: Session,
    school_id: int,
    period_month: int | None = None,
    period_year: int | None = None
):
    query = db.query(Invoice).filter(
        Invoice.school_id == school_id,
        Invoice.status == "approved"
    )

    if period_month:
        query = query.filter(
            Invoice.period_month == period_month
        )

    if period_year:
        query = query.filter(
            Invoice.period_year == period_year
        )

    return query


@router.get("/", response_class=HTMLResponse)
async def payments_dashboard(
    school_id: int,
    request: Request,
    period_month: str | None = None,
    period_year: str | None = None,
    reminder_status: str | None = None,
    reminder_message: str | None = None,
    payment_status: str | None = None,
    payment_message: str | None = None,
    db: Session = Depends(get_db)
):
    period_month = parse_optional_int(period_month)
    period_year = parse_optional_int(period_year)

    user = require_school_access(
        request=request,
        db=db,
        school_id=school_id
    )

    if not user:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    school = db.query(School).filter(
        School.id == school_id
    ).first()

    if not school:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    today = date.today()

    invoices = (
        build_payments_query(
            db=db,
            school_id=school_id,
            period_month=period_month,
            period_year=period_year
        )
        .order_by(Invoice.created_at.desc())
        .all()
    )

    total_invoiced = sum(
        float(invoice.amount or 0)
        for invoice in invoices
    )

    total_paid = sum(
        float(invoice.amount or 0)
        for invoice in invoices
        if invoice.is_paid
    )

    total_pending = total_invoiced - total_paid

    overdue_invoices = [
        invoice for invoice in invoices
        if not invoice.is_paid
        and invoice.due_date
        and invoice.due_date < today
    ]

    pending_invoices = [
        invoice for invoice in invoices
        if not invoice.is_paid
    ]

    total_overdue = sum(
        float(invoice.amount or 0)
        for invoice in overdue_invoices
    )

    paid_count = len([
        invoice for invoice in invoices
        if invoice.is_paid
    ])

    pending_count = len(pending_invoices)
    overdue_count = len(overdue_invoices)

    paid_percentage = 0

    if total_invoiced > 0:
        paid_percentage = round(
            (total_paid / total_invoiced) * 100,
            2
        )

    overdue_student_ids = {
        invoice.student_id
        for invoice in overdue_invoices
        if invoice.student_id
    }

    students_by_id = {
        student.id: student
        for student in db.query(Student)
        .filter(Student.school_id == school_id)
        .all()
    }

    return templates.TemplateResponse(
        "payments/dashboard.html",
        {
            "request": request,
            "school": school,
            "period_month": period_month,
            "period_year": period_year,
            "reminder_status": reminder_status,
            "reminder_message": reminder_message,
            "payment_status": payment_status,
            "payment_message": payment_message,
            "total_invoiced": total_invoiced,
            "total_paid": total_paid,
            "total_pending": total_pending,
            "total_overdue": total_overdue,
            "paid_count": paid_count,
            "pending_count": pending_count,
            "overdue_count": overdue_count,
            "overdue_students_count": len(overdue_student_ids),
            "paid_percentage": paid_percentage,
            "pending_invoices": pending_invoices,
            "overdue_invoices": overdue_invoices,
            "students_by_id": students_by_id,
            "today": today
        }
    )


@router.post("/{invoice_id}/pay")
async def register_payment(
    school_id: int,
    invoice_id: int,
    request: Request,
    payment_method: str = Form(...),
    payment_note: str = Form(""),
    db: Session = Depends(get_db)
):
    user = require_school_access(
        request=request,
        db=db,
        school_id=school_id
    )

    if not user:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.school_id == school_id,
        Invoice.status == "approved"
    ).first()

    if not invoice:
        return RedirectResponse(
            url=f"/schools/{school_id}/payments?payment_status=error&payment_message=Factura no encontrada",
            status_code=303
        )

    if invoice.is_paid:
        return RedirectResponse(
            url=f"/schools/{school_id}/payments?payment_status=error&payment_message=La factura ya estaba marcada como pagada",
            status_code=303
        )

    invoice.is_paid = True
    invoice.paid_at = datetime.now(timezone.utc)
    invoice.payment_method = payment_method
    invoice.payment_note = payment_note

    db.commit()

    InvoiceLogService.create(
        db=db,
        school_id=school_id,
        student_id=invoice.student_id,
        invoice_id=invoice.id,
        event_type="payment_registered",
        message=f"Pago registrado desde cobranzas. Método: {payment_method}. Nota: {payment_note or '-'}"
    )

    return RedirectResponse(
        url=f"/schools/{school_id}/payments?payment_status=success&payment_message=Pago registrado correctamente",
        status_code=303
    )


@router.post("/{invoice_id}/send-reminder")
async def send_payment_reminder(
    school_id: int,
    invoice_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(
        request=request,
        db=db,
        school_id=school_id
    )

    if not user:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    school = db.query(School).filter(
        School.id == school_id
    ).first()

    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.school_id == school_id,
        Invoice.status == "approved"
    ).first()

    if not school or not invoice:
        return RedirectResponse(
            url=f"/schools/{school_id}/payments",
            status_code=303
        )

    if invoice.is_paid:
        return RedirectResponse(
            url=f"/schools/{school_id}/payments?reminder_status=error&reminder_message=La factura ya se encuentra pagada",
            status_code=303
        )

    student = db.query(Student).filter(
        Student.id == invoice.student_id,
        Student.school_id == school_id
    ).first()

    settings = db.query(SchoolSettings).filter(
        SchoolSettings.school_id == school_id
    ).first()

    if not student:
        return RedirectResponse(
            url=f"/schools/{school_id}/payments?reminder_status=error&reminder_message=Alumno no encontrado",
            status_code=303
        )

    if not settings:
        return RedirectResponse(
            url=f"/schools/{school_id}/payments?reminder_status=error&reminder_message=Configuración SMTP no encontrada",
            status_code=303
        )

    today = date.today()

    period = "-"

    if invoice.period_month and invoice.period_year:
        period = f"{invoice.period_month}/{invoice.period_year}"

    invoice_number = get_invoice_number(invoice)

    email_result = EmailService.send_payment_reminder_email(
        smtp_host=settings.smtp_host,
        smtp_port=settings.smtp_port or 587,
        smtp_user=settings.smtp_user,
        smtp_password=settings.smtp_password,
        smtp_from=settings.smtp_from,
        to_email=student.email,
        student_name=student.full_name,
        school_name=school.name,
        period=period,
        amount=float(invoice.amount or 0),
        due_date=str(invoice.due_date or ""),
        invoice_number=invoice_number,
        is_overdue=bool(invoice.due_date and invoice.due_date < today)
    )

    if email_result["sent"]:
        InvoiceLogService.create(
            db=db,
            school_id=school_id,
            student_id=student.id,
            invoice_id=invoice.id,
            event_type="payment_reminder_sent",
            message=f"Recordatorio de pago enviado a {student.email}."
        )

        return RedirectResponse(
            url=f"/schools/{school_id}/payments?reminder_status=success&reminder_message=Recordatorio enviado correctamente",
            status_code=303
        )

    InvoiceLogService.create(
        db=db,
        school_id=school_id,
        student_id=student.id,
        invoice_id=invoice.id,
        event_type="payment_reminder_error",
        message=f"Error enviando recordatorio: {email_result['error']}"
    )

    return RedirectResponse(
        url=f"/schools/{school_id}/payments?reminder_status=error&reminder_message={email_result['error']}",
        status_code=303
    )


@router.get("/export")
async def export_payments_excel(
    school_id: int,
    request: Request,
    period_month: str | None = None,
    period_year: str | None = None,
    db: Session = Depends(get_db)
):
    period_month = parse_optional_int(period_month)
    period_year = parse_optional_int(period_year)

    user = require_school_access(
        request=request,
        db=db,
        school_id=school_id
    )

    if not user:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    school = db.query(School).filter(
        School.id == school_id
    ).first()

    if not school:
        return RedirectResponse(
            url="/login",
            status_code=303
        )

    today = date.today()

    invoices = (
        build_payments_query(
            db=db,
            school_id=school_id,
            period_month=period_month,
            period_year=period_year
        )
        .order_by(Invoice.created_at.desc())
        .all()
    )

    students_by_id = {
        student.id: student
        for student in db.query(Student)
        .filter(Student.school_id == school_id)
        .all()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Cobranzas"

    headers = [
        "Alumno",
        "DNI Alumno",
        "Tutor",
        "DNI Tutor",
        "Email",
        "Periodo",
        "Factura",
        "Importe",
        "Fecha Vencimiento",
        "Pagada",
        "Fecha Pago",
        "Método de Pago",
        "Nota de Pago",
        "Estado"
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color="DDDDDD",
        end_color="DDDDDD",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for invoice in invoices:
        student = students_by_id.get(invoice.student_id)

        if invoice.is_paid:
            payment_status = "Pagada"
        elif invoice.due_date and invoice.due_date < today:
            payment_status = "Vencida"
        else:
            payment_status = "Pendiente"

        periodo = "-"

        if invoice.period_month and invoice.period_year:
            periodo = f"{invoice.period_month}/{invoice.period_year}"

        ws.append([
            student.full_name if student else "",
            student.dni if student else "",
            student.guardian_name if student else "",
            student.guardian_dni if student else "",
            student.email if student else "",
            periodo,
            get_invoice_number(invoice),
            float(invoice.amount or 0),
            str(invoice.due_date or ""),
            "SI" if invoice.is_paid else "NO",
            str(invoice.paid_at or ""),
            invoice.payment_method or "",
            invoice.payment_note or "",
            payment_status
        ])

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = str(cell.value or "")
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(
            max_length + 2,
            40
        )

    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    )

    wb.save(temp_file.name)

    filename = "cobranzas.xlsx"

    if period_month and period_year:
        filename = f"cobranzas_{period_month}_{period_year}.xlsx"

    return FileResponse(
        temp_file.name,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )