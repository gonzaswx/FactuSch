import tempfile

from fastapi import APIRouter
from fastapi import Depends
from fastapi import File
from fastapi import Form
from fastapi import Request
from fastapi import UploadFile

from fastapi.responses import FileResponse
from fastapi.responses import HTMLResponse
from fastapi.responses import RedirectResponse

from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from database import get_db

from models.school import School
from models.school_level import SchoolLevel
from models.student import Student

from services.student_service import StudentService
from services.student_import_service import StudentImportService

from utils.templates import templates

from routers.auth import require_school_access


router = APIRouter(
    prefix="/schools/{school_id}/students",
    tags=["Students"]
)


def get_school_or_redirect(db: Session, school_id: int):
    return db.query(School).filter(
        School.id == school_id
    ).first()


def get_active_levels(db: Session, school_id: int):
    return (
        db.query(SchoolLevel)
        .filter(
            SchoolLevel.school_id == school_id,
            SchoolLevel.is_active == True
        )
        .order_by(SchoolLevel.name.asc())
        .all()
    )


def parse_optional_int(value):
    if value is None or value == "":
        return None

    try:
        return int(value)
    except ValueError:
        return None


def build_student_form(
    full_name="",
    dni="",
    email="",
    course="",
    division="",
    guardian_name="",
    guardian_dni="",
    level_id=None,
    monthly_fee=0
):
    return {
        "full_name": full_name,
        "dni": dni,
        "email": email,
        "course": course,
        "division": division,
        "guardian_name": guardian_name,
        "guardian_dni": guardian_dni,
        "level_id": level_id,
        "monthly_fee": monthly_fee
    }


@router.get("/", response_class=HTMLResponse)
async def list_students(
    school_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    if not school:
        return RedirectResponse(url="/login", status_code=303)

    students = StudentService.get_by_school(db, school_id)

    levels = (
        db.query(SchoolLevel)
        .filter(SchoolLevel.school_id == school_id)
        .all()
    )

    levels_by_id = {
        level.id: level
        for level in levels
    }

    return templates.TemplateResponse(
        "students/list.html",
        {
            "request": request,
            "school": school,
            "students": students,
            "levels_by_id": levels_by_id
        }
    )


@router.get("/create", response_class=HTMLResponse)
async def create_student_page(
    school_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    if not school:
        return RedirectResponse(url="/login", status_code=303)

    return templates.TemplateResponse(
        "students/create.html",
        {
            "request": request,
            "school": school,
            "levels": get_active_levels(db, school_id),
            "error": None,
            "form": build_student_form()
        }
    )


@router.post("/create", response_class=HTMLResponse)
async def create_student(
    school_id: int,
    request: Request,
    full_name: str = Form(...),
    dni: str = Form(...),
    email: str = Form(None),
    course: str = Form(None),
    division: str = Form(None),
    guardian_name: str = Form(None),
    guardian_dni: str = Form(None),
    level_id: str | None = Form(None),
    monthly_fee: float | None = Form(0),
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    if not school:
        return RedirectResponse(url="/login", status_code=303)

    clean_level_id = parse_optional_int(level_id)

    form = build_student_form(
        full_name=full_name,
        dni=dni,
        email=email,
        course=course,
        division=division,
        guardian_name=guardian_name,
        guardian_dni=guardian_dni,
        level_id=clean_level_id,
        monthly_fee=monthly_fee
    )

    try:
        StudentService.create(
            db=db,
            school_id=school_id,
            full_name=full_name,
            dni=dni,
            email=email,
            course=course,
            division=division,
            guardian_name=guardian_name,
            guardian_dni=guardian_dni,
            level_id=clean_level_id,
            monthly_fee=monthly_fee
        )

    except Exception as e:
        return templates.TemplateResponse(
            "students/create.html",
            {
                "request": request,
                "school": school,
                "levels": get_active_levels(db, school_id),
                "error": str(e),
                "form": form
            },
            status_code=400
        )

    return RedirectResponse(
        url=f"/schools/{school_id}/students",
        status_code=303
    )


@router.get("/{student_id}/edit", response_class=HTMLResponse)
async def edit_student_page(
    school_id: int,
    student_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    student = StudentService.get_by_id(
        db=db,
        school_id=school_id,
        student_id=student_id
    )

    if not school or not student:
        return RedirectResponse(
            url=f"/schools/{school_id}/students",
            status_code=303
        )

    return templates.TemplateResponse(
        "students/edit.html",
        {
            "request": request,
            "school": school,
            "student": student,
            "levels": get_active_levels(db, school_id),
            "error": None
        }
    )


@router.post("/{student_id}/edit", response_class=HTMLResponse)
async def edit_student(
    school_id: int,
    student_id: int,
    request: Request,
    full_name: str = Form(...),
    dni: str = Form(...),
    email: str = Form(None),
    course: str = Form(None),
    division: str = Form(None),
    guardian_name: str = Form(None),
    guardian_dni: str = Form(None),
    level_id: str | None = Form(None),
    monthly_fee: float | None = Form(0),
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    clean_level_id = parse_optional_int(level_id)

    try:
        StudentService.update(
            db=db,
            school_id=school_id,
            student_id=student_id,
            full_name=full_name,
            dni=dni,
            email=email,
            course=course,
            division=division,
            guardian_name=guardian_name,
            guardian_dni=guardian_dni,
            level_id=clean_level_id,
            monthly_fee=monthly_fee
        )

    except Exception as e:
        school = get_school_or_redirect(db, school_id)

        student = StudentService.get_by_id(
            db=db,
            school_id=school_id,
            student_id=student_id
        )

        return templates.TemplateResponse(
            "students/edit.html",
            {
                "request": request,
                "school": school,
                "student": student,
                "levels": get_active_levels(db, school_id),
                "error": str(e)
            },
            status_code=400
        )

    return RedirectResponse(
        url=f"/schools/{school_id}/students",
        status_code=303
    )


@router.post("/{student_id}/toggle-active")
async def toggle_student_active(
    school_id: int,
    student_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    student = db.query(Student).filter(
        Student.id == student_id,
        Student.school_id == school_id
    ).first()

    if student:
        student.is_active = not student.is_active
        db.commit()

    return RedirectResponse(
        url=f"/schools/{school_id}/students",
        status_code=303
    )


@router.get("/import", response_class=HTMLResponse)
async def import_students_page(
    school_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    if not school:
        return RedirectResponse(url="/login", status_code=303)

    levels = get_active_levels(db, school_id)

    return templates.TemplateResponse(
        "students/import.html",
        {
            "request": request,
            "school": school,
            "levels": levels,
            "result": None
        }
    )


@router.get("/import/template")
async def download_students_template(
    school_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    if not school:
        return RedirectResponse(url="/login", status_code=303)

    levels = get_active_levels(db, school_id)

    wb = Workbook()

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    required_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )

    # Hoja 1: Alumnos
    ws = wb.active
    ws.title = "Alumnos"

    headers = [
        "full_name",
        "dni",
        "email",
        "course",
        "division",
        "guardian_name",
        "guardian_dni",
        "nivel_alias",
        "monthly_fee"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for cell in ["A1", "B1", "H1"]:
        ws[cell].fill = required_fill

    ws.append([
        "Juan Perez",
        "40111222",
        "responsable@mail.com",
        "5",
        "A",
        "Carlos Perez",
        "20111222",
        levels[0].alias if levels else "PRIMARIO",
        ""
    ])

    ws.append([
        "Maria Lopez",
        "41222333",
        "responsable2@mail.com",
        "4",
        "B",
        "Ana Lopez",
        "22111222",
        levels[1].alias if len(levels) > 1 else "SECUNDARIO",
        "50000"
    ])

    # Hoja 2: Niveles disponibles
    ws_levels = wb.create_sheet("Niveles disponibles")
    ws_levels.append(["Nivel", "Alias para Excel", "Cuota mensual"])

    for cell in ws_levels[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for level in levels:
        ws_levels.append([
            level.name,
            level.alias,
            float(level.monthly_fee or 0)
        ])

    # Hoja 3: Instrucciones
    ws_info = wb.create_sheet("Instrucciones")
    ws_info.append(["Campo", "Obligatorio", "Descripción"])

    for cell in ws_info[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    instructions = [
        ["full_name", "SI", "Nombre completo del alumno."],
        ["dni", "SI", "DNI del alumno sin puntos ni espacios."],
        ["email", "NO", "Email del responsable o tutor."],
        ["course", "NO", "Curso o año del alumno."],
        ["division", "NO", "División del alumno. Ejemplo: A, B, C."],
        ["guardian_name", "NO", "Nombre del responsable o tutor."],
        ["guardian_dni", "NO", "DNI del tutor. No puede ser igual al DNI del alumno."],
        ["nivel_alias", "SI", "Alias del nivel. Debe existir previamente en Niveles."],
        ["monthly_fee", "NO", "Si queda vacío o en 0, se usa la cuota mensual del nivel."]
    ]

    for row in instructions:
        ws_info.append(row)

    ws_info.append([])
    ws_info.append(["Reglas importantes"])
    ws_info.append(["1", "Crear primero los niveles en FactuSch."])
    ws_info.append(["2", "Usar el alias exacto informado en la hoja Niveles disponibles."])
    ws_info.append(["3", "Las mayúsculas/minúsculas no importan."])
    ws_info.append(["4", "Si una fila tiene error, se omite esa fila y se importan las demás."])
    ws_info.append(["5", "No modificar los nombres de las columnas."])

    for worksheet in [ws, ws_levels, ws_info]:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = str(cell.value or "")
                max_length = max(max_length, len(value))

            worksheet.column_dimensions[column_letter].width = min(
                max_length + 4,
                55
            )

    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    )

    wb.save(temp_file.name)

    return FileResponse(
        temp_file.name,
        filename=f"plantilla_alumnos_{school.id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/import", response_class=HTMLResponse)
async def import_students(
    school_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    user = require_school_access(request, db, school_id)

    if not user:
        return RedirectResponse(url="/login", status_code=303)

    school = get_school_or_redirect(db, school_id)

    if not school:
        return RedirectResponse(url="/login", status_code=303)

    content = await file.read()

    result = StudentImportService.import_file(
        db=db,
        school_id=school_id,
        file_content=content,
        filename=file.filename
    )

    levels = get_active_levels(db, school_id)

    return templates.TemplateResponse(
        "students/import.html",
        {
            "request": request,
            "school": school,
            "levels": levels,
            "result": result
        }
    )